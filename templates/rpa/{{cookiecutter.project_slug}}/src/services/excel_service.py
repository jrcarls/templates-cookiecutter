import logging
from pathlib import Path
import openpyxl

logger = logging.getLogger(__name__)


class ExcelService:
    def read(self, path: Path) -> list[dict]:
        logger.info("Lendo planilha: %s", path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        return [
            dict(zip(headers, (cell.value for cell in row)))
            for row in ws.iter_rows(min_row=2)
        ]

    def write(self, path: Path, rows: list[dict]) -> None:
        logger.info("Escrevendo planilha: %s", path)
        wb = openpyxl.Workbook()
        ws = wb.active
        if not rows:
            wb.save(path)
            return
        ws.append(list(rows[0].keys()))
        for row in rows:
            ws.append(list(row.values()))
        wb.save(path)
